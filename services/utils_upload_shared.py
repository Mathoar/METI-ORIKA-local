import sqlite3
import pandas as pd
import re
import unicodedata
import logging
from openpyxl import load_workbook
from services.utils import DB_NAME



logger = logging.getLogger(__name__)

def reset_meti_table():
    try:
        with sqlite3.connect(DB_NAME) as conn:
            conn.execute("DELETE FROM meti")
            conn.commit()
            logger.info("Table 'meti' cleared.")
        return True
    except Exception as e:
        logger.error(f"Error resetting meti table: {e}")
        return False

def reset_rupture_meti_table():
    try:
        with sqlite3.connect(DB_NAME) as conn:
            conn.execute("DELETE FROM rupture_meti")
            conn.commit()
            logger.info("Table 'rupture_meti' cleared.")
        return True
    except Exception as e:
        logger.error(f"Error resetting rupture_meti table: {e}")
        return False

def normalize_column(col):
    if not isinstance(col, str):
        return str(col).upper()
    col = ''.join(c for c in unicodedata.normalize('NFD', col) if unicodedata.category(c) != 'Mn')
    col = col.replace(' ', '').replace('-', '').replace('.', '').replace('°', '').upper()
    return col

def clean_currency(value):
    if pd.isna(value) or str(value).strip() == '':
        return 0.0
    cleaned = re.sub(r'[€,\s]', '', str(value))
    try:
        return float(cleaned)
    except ValueError:
        logger.warning(f"Invalid currency value: {value}")
        return 0.0

def clean_meti_excel(input_path, output_path, header_row, data_row):
    try:
        wb = load_workbook(input_path)
        ws = wb.active

        if header_row < 1 or data_row <= header_row:
            logger.error(f"Invalid row numbers: header_row={header_row}, data_row={data_row}")
            return False

        if header_row > 1:
            ws.delete_rows(1, header_row - 1)

        headers = [cell.value for cell in ws[1]]
        required_headers = ['Nomenclature - Article', 'CA HT', 'Passage', 'Marge', 'PM', 'TdM', 'Poids promo']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            logger.error(f"Missing required headers: {missing_headers}")
            return False

        ca_ht_col_idx = headers.index("CA HT") + 1
        adjusted_data_row = data_row - (header_row - 1)
        if adjusted_data_row < 2:
            logger.error(f"Invalid data row after adjustment: {adjusted_data_row}")
            return False

        for cell in ws[adjusted_data_row]:
            if cell.value in (None, "", " "):
                cell.value = "Inconnu"

        for row in ws.iter_rows(min_row=adjusted_data_row + 1, max_row=ws.max_row, min_col=ca_ht_col_idx, max_col=ca_ht_col_idx):
            if row[0].value in (None, "", " "):
                row[0].value = 0

        wb.save(output_path)
        return True
    except Exception as e:
        logger.error(f"Error cleaning METI file: {e}")
        return False

def insert_meti_data(df):
    try:
        with sqlite3.connect(DB_NAME) as conn:
            cursor = conn.cursor()
            df.columns = [normalize_column(col) for col in df.columns]
            required_columns = ['NOMENCLATUREARTICLE', 'CAHT', 'PASSAGE', 'MARGE', 'PM', 'TDM', 'POIDSPROMO']
            missing = [col for col in required_columns if col not in df.columns]
            if missing:
                raise ValueError(f"Missing required columns: {missing}")

            df['NOMENCLATUREARTICLE'] = df['NOMENCLATUREARTICLE'].fillna('Unknown').astype(str)
            df['CAHT'] = df['CAHT'].apply(clean_currency)
            df['PASSAGE'] = df['PASSAGE'].fillna(0).astype(float)
            df['MARGE'] = df['MARGE'].apply(clean_currency)
            df['PM'] = df['PM'].apply(clean_currency)
            df['TDM'] = df['TDM'].fillna(0.0).astype(float)
            df['POIDSPROMO'] = df['POIDSPROMO'].fillna(0.0).astype(float)

            def split_nomenclature(val):
                if isinstance(val, str) and '-' in val:
                    parts = val.split('-', 1)
                    return parts[0].strip(), parts[1].strip().upper()
                return '', val.strip().upper() if isinstance(val, str) else ''

            df[['GENERATED_ID', 'GENERATED_ARTICLE']] = df['NOMENCLATUREARTICLE'].apply(lambda x: pd.Series(split_nomenclature(x)))
            df = df[['NOMENCLATUREARTICLE', 'CAHT', 'PASSAGE', 'MARGE', 'PM', 'TDM', 'POIDSPROMO', 'GENERATED_ARTICLE', 'GENERATED_ID']]

            df_valid = df[df['CAHT'].notnull() & (df['CAHT'] > 0)]
            df_rupture = df[~(df['CAHT'].notnull() & (df['CAHT'] > 0))]

            for _, row in df_valid.iterrows():
                cursor.execute('''INSERT OR REPLACE INTO meti (
                    nomenclature, ca_ht, passage, marge, pm, tdm, poids_promo, generated_article, generated_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', tuple(row))

            for _, row in df_rupture.iterrows():
                cursor.execute('''INSERT INTO rupture_meti (
                    nomenclature, ca_ht, passage, marge, pm, tdm, poids_promo, generated_article, generated_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', tuple(row))

            conn.commit()
        return True
    except Exception as e:
        logger.error(f"Error inserting METI data: {e}")
        return False
